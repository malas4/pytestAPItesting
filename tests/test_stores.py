import requests
import openpyxl
from config import BASE_URI

# data driven testing by getting the data from excel sheet
def test_number_of_stores_in_zipcode():

    f=openpyxl.load_workbook(r'C:\Users\Nihaan\PycharmProjects\rest\bestbuyzips.xlsx')
    sheet = f['Sheet1']
    rows = sheet.max_row
    for i in range(1, rows):
        zip_code = sheet.cell(row=i+1, column=1).value
        try:
            response = requests.get(f'{BASE_URI}/stores?zip={zip_code}')
            response.raise_for_status()
        except requests.exceptions.HTTPError as e:
            print(e)
        response_json = response.json()
        no_of_stores_in_response = response_json['total']
        if no_of_stores_in_response  == sheet.cell(row=i+1, column=2).value:
            assert True
        else:
            print(f'mismatch in no of stores for zipcode {zip_code}')
            assert False




